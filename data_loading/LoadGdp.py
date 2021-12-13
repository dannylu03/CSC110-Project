"""CSC110 Final Project: Data Loading: Load Gdp

Copyright and Usage Information
===============================

This file is Copyright (c) 2021 Clark Zhang, Danny Lu, Alex Balaria, Yue Fung Lee.
"""
from data_loading.DataLoad import DataLoad
import openpyxl


class LoadGdp(DataLoad):
    """Load Gdp data"""

    def __init__(self, file_path: str):
        super().__init__(file_path)
        self.raw_data = []

    def load_data(self) -> list[(str, float)]:
        """ Loads data from provided file path


         """
        sheets = openpyxl.open(self.file_path)
        sheet_data = sheets['Data']
        i = 6
        # reads cells while they contain relevant values
        while sheet_data.cell(i, 2).value is not None:
            self.raw_data.append((sheet_data.cell(i, 2).value, float(sheet_data.cell(i, 3).value)))
            i += 1

        return self.raw_data
